import discord
from discord import app_commands
from discord.ext import commands, tasks
import openpyxl
import io
import os
import json
import asyncio
import calendar
import logging
from datetime import datetime, timezone
from dotenv import load_dotenv

# ─────────────────────────────────────────────
#  LOGGING SETUP
# ─────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("monsterbot.log", encoding="utf-8"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger("monsterbot")

load_dotenv()

# ─────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────
def get_env_int(key):
    val = os.getenv(key, "0").strip()
    try:
        return int(val)
    except ValueError:
        log.error(f"Ongeldige waarde voor {key}: '{val}'")
        return 0

TOKEN              = os.getenv("DISCORD_TOKEN", "").strip()
GUILD_ID           = get_env_int("GUILD_ID")
TICKET_CAT_ID       = get_env_int("TICKET_CATEGORY_ID")
TICKET_BUTTON_CH_ID = get_env_int("TICKET_BUTTON_CHANNEL_ID")
STAFF_ROLE_ID       = get_env_int("STAFF_ROLE_ID")
LOG_CHANNEL_ID     = get_env_int("LOG_CHANNEL_ID")
LEADERBOARD_CH_ID  = get_env_int("LEADERBOARD_CHANNEL_ID")
PAYMENT_CH_ID      = get_env_int("PAYMENT_CHANNEL_ID")
TEMPLATE_PATH      = os.path.join(os.path.dirname(__file__), "Clipfarming_Verdiensten_2026.xlsx")
DATA_FILE          = os.path.join(os.path.dirname(__file__), "monthly_data.json")
STATE_FILE         = os.path.join(os.path.dirname(__file__), "bot_state.json")

OPEN_DAYS = {1, 2, 3, 4}

MONTHS_NL = [
    "Januari","Februari","Maart","April","Mei","Juni",
    "Juli","Augustus","September","Oktober","November","December"
]

if not TOKEN:
    log.critical("DISCORD_TOKEN ontbreekt in .env - bot kan niet starten")
    raise SystemExit(1)

# ─────────────────────────────────────────────
#  INTENTS & BOT
# ─────────────────────────────────────────────
intents = discord.Intents.default()
intents.message_content = True
intents.members = True

bot  = commands.Bot(command_prefix="!", intents=intents)
tree = bot.tree

# ─────────────────────────────────────────────
#  WACHTENDE BETAALLINK-VERZOEKEN
#  { user_id: { "bedrag": float, "naam": str, "ticket_channel_id": int, "guild_id": int } }
# ─────────────────────────────────────────────
pending_payment_links: dict[int, dict] = {}

# ─────────────────────────────────────────────
#  HELPERS — tijd
# ─────────────────────────────────────────────
def now_utc() -> datetime:
    return datetime.now(timezone.utc)

def now_local() -> datetime:
    return datetime.now()

# ─────────────────────────────────────────────
#  STATE
# ─────────────────────────────────────────────
def load_state():
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return {"bot_enabled": True}

def save_state(state):
    try:
        with open(STATE_FILE, "w") as f:
            json.dump(state, f, indent=2)
    except Exception as e:
        log.error(f"Kon state niet opslaan: {e}")

def is_bot_open():
    state = load_state()
    if not state.get("bot_enabled", True):
        return False
    return now_local().day in OPEN_DAYS

# ─────────────────────────────────────────────
#  LEADERBOARD DATA
# ─────────────────────────────────────────────
def calc_earnings(views):
    if views >= 1_000_000:
        return 250.0
    return float(views // 10_000)

def load_data():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r") as f:
                return json.load(f)
        except Exception as e:
            log.error(f"Kon data niet laden: {e}")
    return {}

def save_data(data):
    try:
        with open(DATA_FILE, "w") as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        log.error(f"Kon data niet opslaan: {e}")

def get_month_key():
    now = now_local()
    return f"{now.year}-{now.month:02d}"

def add_to_leaderboard(discord_naam, user_id, views, earnings):
    data = load_data()
    key  = get_month_key()
    if key not in data:
        data[key] = {}
    uid = str(user_id)
    if uid not in data[key]:
        data[key][uid] = {"naam": discord_naam, "views": 0, "earnings": 0.0}
    data[key][uid]["views"]    += max(0, int(views))
    data[key][uid]["earnings"] += max(0.0, float(earnings))
    data[key][uid]["naam"]      = str(discord_naam)[:50]
    save_data(data)

# ─────────────────────────────────────────────
#  HELPERS — overig
# ─────────────────────────────────────────────
async def send_log(guild, embed, file=None):
    if not LOG_CHANNEL_ID:
        return
    try:
        ch = guild.get_channel(LOG_CHANNEL_ID)
        if not ch:
            return
        await (ch.send(embed=embed, file=file) if file else ch.send(embed=embed))
    except discord.Forbidden:
        log.warning("Geen toegang tot log-kanaal")
    except Exception as e:
        log.error(f"send_log fout: {e}")

async def send_dm(user, embed, view=None):
    try:
        if view:
            await user.send(embed=embed, view=view)
        else:
            await user.send(embed=embed)
        return True
    except discord.Forbidden:
        log.warning(f"Kon geen DM sturen naar {user} — DMs uitgeschakeld")
        return False
    except Exception as e:
        log.error(f"send_dm fout voor {user}: {e}")
        return False

def is_staff(member):
    if not STAFF_ROLE_ID:
        return False
    role = member.guild.get_role(STAFF_ROLE_ID)
    return role in member.roles if role else False

def sanitize(text, max_len=200):
    if not text:
        return ""
    clean = str(text).strip()
    clean = clean.replace("@everyone", "").replace("@here", "")
    return clean[:max_len]

# ─────────────────────────────────────────────
#  EXCEL PARSING
# ─────────────────────────────────────────────
def parse_submission(file_bytes):
    if len(file_bytes) > 5 * 1024 * 1024:
        log.warning("Ingediend bestand te groot (>5MB)")
        return None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        log.warning(f"Excel kon niet worden geopend: {e}")
        return None

    results      = []
    total_views  = 0
    total        = 0.0
    discord_naam = None
    email        = None

    for sheet_name in wb.sheetnames[:12]:
        try:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=4, max_row=200, max_col=6):
                naam_cell    = row[0].value if len(row) > 0 else None
                email_cell   = row[1].value if len(row) > 1 else None
                platform_val = row[2].value if len(row) > 2 else None
                link_val     = row[3].value if len(row) > 3 else None
                views_val    = row[4].value if len(row) > 4 else None

                if not views_val or str(views_val).strip() in ("", "Views (werkelijk)", "Views"):
                    continue
                views_str = str(views_val).strip()
                if views_str.startswith("[") or views_str.startswith("="):
                    continue

                try:
                    if isinstance(views_val, (int, float)):
                        views = int(views_val)
                    else:
                        clean_val = str(views_val).strip().replace(" ", "")
                        if "," in clean_val and "." in clean_val:
                            clean_val = clean_val.replace(".", "").replace(",", ".")
                        elif "," in clean_val:
                            clean_val = clean_val.replace(",", ".")
                        views = int(float(clean_val))
                    if views <= 0 or views > 100_000_000:
                        continue
                except (ValueError, OverflowError, TypeError):
                    continue

                if not discord_naam and naam_cell:
                    naam_str = str(naam_cell).strip()
                    if "[Vul" not in naam_str and naam_str:
                        discord_naam = sanitize(naam_str, 50)
                if not email and email_cell:
                    email_str = str(email_cell).strip()
                    if "[Vul" not in email_str and email_str:
                        email = sanitize(email_str, 100)

                floored = (views // 10_000) * 10_000
                earning = calc_earnings(floored)
                total       += earning
                total_views += floored

                link = sanitize(link_val, 200) if link_val else ""
                if link and not link.startswith(("http://", "https://")):
                    link = "Ongeldige link"

                results.append({
                    "platform": sanitize(platform_val, 30) if platform_val else "Onbekend",
                    "link":     link,
                    "views":    views,
                    "floored":  floored,
                    "earning":  earning,
                })

                if len(results) >= 20:
                    break
            if len(results) >= 20:
                break
        except Exception as e:
            log.warning(f"Fout bij lezen sheet {sheet_name}: {e}")
            continue

    if not results:
        return None

    return {
        "discord_naam": discord_naam or "Onbekend",
        "email":        email or "Onbekend",
        "clips":        results,
        "total_views":  total_views,
        "total":        round(total, 2),
    }

# ─────────────────────────────────────────────
#  EMBEDS
# ─────────────────────────────────────────────
def build_summary_embed(data, user):
    embed = discord.Embed(
        title="Betalingsoverzicht - Ingediend",
        description=(
            f"**Discord:** {sanitize(data['discord_naam'])}\n"
            f"**E-mail:** {sanitize(data['email'])}\n"
            f"**Ingediend door:** {user.mention}\n"
            f"**Datum:** {now_local().strftime('%d %B %Y')}"
        ),
        color=0x7b2ff7,
        timestamp=now_utc(),
    )
    clips = data["clips"]
    current_field_content = ""
    field_count = 1

    for i, c in enumerate(clips[:20], 1):
        afgerond = f" *(afgerond naar {c['floored']:,})*" if c["views"] != c["floored"] else ""
        link_display = c['link'][:40] + ('...' if len(c['link']) > 40 else '')
        line = (
            f"`{i}.` **{c['platform']}** | {c['views']:,} views{afgerond} - EUR {c['earning']:.2f}\n"
            f"   {link_display}\n"
        )
        if len(current_field_content) + len(line) > 1000:
            embed.add_field(name=f"Clips Deel {field_count}", value=current_field_content, inline=False)
            current_field_content = line
            field_count += 1
        else:
            current_field_content += line

    if current_field_content:
        field_name = f"Clips ({len(clips)})" if field_count == 1 else f"Clips Deel {field_count}"
        embed.add_field(name=field_name, value=current_field_content, inline=False)
    elif not clips:
        embed.add_field(name="Clips", value="Geen clips", inline=False)

    embed.add_field(name="Totaal te betalen", value=f"**EUR {data['total']:.2f}**", inline=False)
    embed.set_footer(text="MONSTERBOT - Wacht op goedkeuring van staff")
    return embed

def build_prijslijst_embed():
    embed = discord.Embed(
        title="Prijslijst",
        description=(
            "EUR 1 per 10.000 views - altijd afgerond naar beneden.\n"
            "Views gelden voor een termijn van 1 maand.\n"
            "Voorbeeld: 16.000 views = 10.000 = EUR 1"
        ),
        color=0x00c9a7,
    )
    embed.add_field(name="Tarieven", value="\n".join([
        "10.000 views → EUR 1",
        "20.000 views → EUR 2",
        "50.000 views → EUR 5",
        "100.000 views → EUR 10",
        "250.000 views → EUR 25",
        "500.000 views → EUR 50",
        "1.000.000 views → EUR 250",
    ]), inline=False)
    return embed

def bot_closed_embed():
    next_month = MONTHS_NL[now_local().month % 12]
    return discord.Embed(
        title="Bot momenteel gesloten",
        description=(
            f"Betalingen kunnen alleen in de **eerste 4 dagen van de maand** worden ingediend.\n\n"
            f"De bot opent weer op **1 {next_month}**."
        ),
        color=0xff4444,
    )

# ─────────────────────────────────────────────
#  BETAALVOORKEUR MODAL
# ─────────────────────────────────────────────

# ─────────────────────────────────────────────
#  TICKET KNOP (gepost in ticket-button kanaal)
# ─────────────────────────────────────────────
class TicketButtonView(discord.ui.View):
    """Persistente knop die in het ticket-aanmaken kanaal staat."""
    def __init__(self):
        super().__init__(timeout=None)

    @discord.ui.button(
        label="📋 Maak betaalticket aan",
        style=discord.ButtonStyle.primary,
        custom_id="open_ticket_button",
    )
    async def open_ticket(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not is_bot_open() and not is_staff(interaction.user):
            await interaction.response.send_message(embed=bot_closed_embed(), ephemeral=True)
            return

        category = interaction.guild.get_channel(TICKET_CAT_ID)
        if not category or not isinstance(category, discord.CategoryChannel):
            await interaction.response.send_message(
                "Ticket-categorie niet gevonden. Vraag een admin.", ephemeral=True
            )
            return

        existing = discord.utils.get(
            category.text_channels,
            name=f"ticket-{interaction.user.name.lower().replace(' ', '-')}",
        )
        if existing:
            await interaction.response.send_message(
                f"Je hebt al een open ticket: {existing.mention}", ephemeral=True
            )
            return

        await interaction.response.send_modal(PaymentModal())


async def post_ticket_button(guild: discord.Guild):
    """Post de ticket-knop in het ingestelde kanaal bij opstarten."""
    if not TICKET_BUTTON_CH_ID:
        log.warning("TICKET_BUTTON_CHANNEL_ID niet ingesteld — knop wordt niet gepost")
        return
    ch = guild.get_channel(TICKET_BUTTON_CH_ID)
    if not ch:
        log.error(f"Ticket-button kanaal ({TICKET_BUTTON_CH_ID}) niet gevonden")
        return

    embed = discord.Embed(
        title="💰 Clipfarming Uitbetaling",
        description=(
            "Wil je je clip-verdiensten uitbetaald krijgen?\n\n"
            "Klik op de knop hieronder om een betaalticket aan te maken.\n"
            "Je ontvangt dan een Excel-bestand dat je invult en terugstuurt.\n\n"
            "⏰ **Let op:** tickets kunnen alleen in de **eerste 4 dagen van de maand** worden aangemaakt."
        ),
        color=0x7b2ff7,
    )
    embed.set_footer(text="MONSTERBOT — MonsterTube Clipfarming")
    await ch.send(embed=embed, view=TicketButtonView())
    log.info(f"Ticket-knop gepost in #{ch.name}")


class PaymentModal(discord.ui.Modal, title="Betaalvoorkeur"):
    betaalmethode = discord.ui.TextInput(
        label="Betaalmethode",
        placeholder="Bijv. PayPal, Tikkie, Bankoverschrijving...",
        required=True,
        max_length=50,
    )
    betaalgegevens = discord.ui.TextInput(
        label="Betaalgegevens (e-mail/IBAN/telefoon)",
        placeholder="Bijv. naam@email.com of NL00BANK0123456789",
        required=True,
        max_length=150,
    )

    async def on_submit(self, interaction: discord.Interaction):
        methode  = sanitize(self.betaalmethode.value, 50)
        gegevens = sanitize(self.betaalgegevens.value, 150)
        await create_ticket(interaction, methode, gegevens)

# ─────────────────────────────────────────────
#  TICKET AANMAKEN
# ─────────────────────────────────────────────
@tree.command(name="uitbetaling", description="Maak een betalingsticket aan om je clip-verdiensten in te dienen.")
async def uitbetaling(interaction: discord.Interaction):
    if not is_bot_open() and not is_staff(interaction.user):
        await interaction.response.send_message(embed=bot_closed_embed(), ephemeral=True)
        return

    guild    = interaction.guild
    user     = interaction.user
    category = guild.get_channel(TICKET_CAT_ID)

    if not category or not isinstance(category, discord.CategoryChannel):
        await interaction.response.send_message("Ticket-categorie niet gevonden. Vraag een admin.", ephemeral=True)
        return

    existing = discord.utils.get(category.text_channels, name=f"ticket-{user.name.lower().replace(' ', '-')}")
    if existing:
        await interaction.response.send_message(f"Je hebt al een open ticket: {existing.mention}", ephemeral=True)
        return

    await interaction.response.send_modal(PaymentModal())

async def create_ticket(interaction: discord.Interaction, betaalmethode: str, betaalgegevens: str):
    await interaction.response.defer(ephemeral=True)
    guild      = interaction.guild
    user       = interaction.user
    category   = guild.get_channel(TICKET_CAT_ID)
    staff_role = guild.get_role(STAFF_ROLE_ID)

    if not category:
        await interaction.followup.send("Fout: categorie niet gevonden.", ephemeral=True)
        return

    try:
        overwrites = {
            guild.default_role: discord.PermissionOverwrite(read_messages=False),
            user:               discord.PermissionOverwrite(read_messages=True, send_messages=True, attach_files=True),
            guild.me:           discord.PermissionOverwrite(read_messages=True, send_messages=True, manage_channels=True),
        }
        if staff_role:
            overwrites[staff_role] = discord.PermissionOverwrite(read_messages=True, send_messages=True)

        channel = await guild.create_text_channel(
            name=f"ticket-{user.name.lower().replace(' ', '-')}",
            category=category,
            overwrites=overwrites,
            topic=f"Ticket:{user.id}|{betaalmethode}:{betaalgegevens}|{now_local().strftime('%d/%m/%Y')}",
        )
    except discord.Forbidden:
        await interaction.followup.send("Fout: geen rechten om kanalen aan te maken.", ephemeral=True)
        return
    except Exception as e:
        await interaction.followup.send("Er ging iets mis bij het aanmaken van je ticket.", ephemeral=True)
        log.error(f"create_ticket fout: {e}")
        return

    month_name = MONTHS_NL[now_local().month - 1]
    try:
        with open(TEMPLATE_PATH, "rb") as f:
            file_data = f.read()
    except FileNotFoundError:
        await interaction.followup.send("Fout: Excel template niet gevonden op de server.", ephemeral=True)
        log.critical(f"Template niet gevonden: {TEMPLATE_PATH}")
        return

    excel_file = discord.File(
        io.BytesIO(file_data),
        filename=f"Clipfarming_Verdiensten_{month_name}_{now_local().year}.xlsx"
    )

    embed = discord.Embed(
        title="Betalingsticket aangemaakt",
        description=(
            f"Hey {user.mention}! Download het Excel-bestand hierboven en volg de stappen:\n\n"
            "**1.** Vul het Excel-bestand volledig in\n"
            "**2.** Upload het ingevulde bestand in dit ticket\n"
            "**3.** Wacht tot een stafflid je bestand heeft gecontroleerd\n\n"
            f"**Betaalmethode:** {betaalmethode}\n"
            f"**Betaalgegevens:** {betaalgegevens}"
        ),
        color=0x7b2ff7,
        timestamp=now_utc(),
    )
    embed.set_footer(text="MONSTERBOT - Ticket systeem")

    try:
        msg = await channel.send(
            content=f"{user.mention}" + (f" {staff_role.mention}" if staff_role else ""),
            file=excel_file,
            embeds=[embed, build_prijslijst_embed()],
            view=TicketCloseView(user.id),
        )
        await msg.pin()
    except Exception as e:
        log.error(f"Fout bij sturen ticketbericht: {e}")

    log_embed = discord.Embed(
        title="Ticket aangemaakt",
        description=(
            f"**Gebruiker:** {user.mention} (`{user}`)\n"
            f"**Kanaal:** {channel.mention}\n"
            f"**Betaalmethode:** {betaalmethode}\n"
            f"**Betaalgegevens:** {betaalgegevens}\n"
            f"**Datum:** {now_local().strftime('%d/%m/%Y %H:%M')}"
        ),
        color=0x7b2ff7,
        timestamp=now_utc(),
    )
    log_embed.set_footer(text="MONSTERBOT - Ticket Log")
    await send_log(guild, log_embed)
    log.info(f"Ticket aangemaakt voor {user} ({user.id})")
    await interaction.followup.send(f"Ticket aangemaakt: {channel.mention}", ephemeral=True)

# ─────────────────────────────────────────────
#  TICKET SLUITEN
# ─────────────────────────────────────────────
class TicketCloseView(discord.ui.View):
    def __init__(self, owner_id):
        super().__init__(timeout=None)
        self.owner_id = owner_id

    @discord.ui.button(label="Ticket sluiten", style=discord.ButtonStyle.danger, custom_id="close_ticket")
    async def close_ticket(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not is_staff(interaction.user) and interaction.user.id != self.owner_id:
            await interaction.response.send_message("Geen toegang.", ephemeral=True)
            return
        await interaction.response.send_message("Ticket wordt gesloten in 5 seconden...")
        await asyncio.sleep(5)
        try:
            await interaction.channel.delete(reason=f"Gesloten door {interaction.user}")
        except Exception as e:
            log.error(f"Kon ticket niet sluiten: {e}")

# ─────────────────────────────────────────────
#  BETALING ONTVANGEN KNOP (voor gebruiker in ticket)
# ─────────────────────────────────────────────
class ConfirmReceiptView(discord.ui.View):
    def __init__(self, user_id: int):
        super().__init__(timeout=None)
        self.user_id = user_id

    @discord.ui.button(
        label="✅ Betaling ontvangen",
        style=discord.ButtonStyle.success,
        custom_id="confirm_payment_received"
    )
    async def confirm(self, interaction: discord.Interaction, button: discord.ui.Button):
        if interaction.user.id != self.user_id:
            await interaction.response.send_message(
                "Alleen de eigenaar van dit ticket kan dit bevestigen.", ephemeral=True
            )
            return

        await interaction.response.send_message(
            "✅ Bedankt voor de bevestiging! Dit ticket wordt over **10 seconden** gesloten."
        )

        log_embed = discord.Embed(
            title="Betaling bevestigd door gebruiker",
            description=f"<@{self.user_id}> heeft de ontvangst van de betaling bevestigd.",
            color=0x00c9a7,
            timestamp=now_utc(),
        )
        log_embed.set_footer(text="MONSTERBOT - Ticket Log")
        await send_log(interaction.guild, log_embed)
        log.info(f"Betaling bevestigd door user {self.user_id}")

        await asyncio.sleep(10)
        try:
            await interaction.channel.delete(reason="Betaling bevestigd door gebruiker")
        except Exception as e:
            log.error(f"Kon ticket niet sluiten na bevestiging: {e}")

# ─────────────────────────────────────────────
#  BETAAL KANAAL KNOP (voor staff)
# ─────────────────────────────────────────────
class PaymentProcessView(discord.ui.View):
    def __init__(self, user_id: int, bedrag: str, discord_naam: str):
        super().__init__(timeout=None)
        self.user_id      = user_id
        self.bedrag       = bedrag
        self.discord_naam = discord_naam

    @discord.ui.button(label="✅ Betaald", style=discord.ButtonStyle.success, emoji="💸", custom_id="mark_as_paid")
    async def mark_as_paid(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not is_staff(interaction.user):
            await interaction.response.send_message("Alleen staff mag dit doen.", ephemeral=True)
            return

        await interaction.response.defer()

        log_embed = discord.Embed(
            title="✅ Betaling Voltooid",
            description=(
                f"**Stafflid:** {interaction.user.mention}\n"
                f"**Ontvanger:** <@{self.user_id}> ({self.discord_naam})\n"
                f"**Bedrag:** {self.bedrag}\n"
                f"**Status:** Overgemaakt."
            ),
            color=0x2ecc71,
            timestamp=now_utc(),
        )
        await send_log(interaction.guild, log_embed)
        await interaction.message.delete()
        log.info(f"Betaling aan {self.discord_naam} gemarkeerd als betaald door {interaction.user}")

# ─────────────────────────────────────────────
#  STAFF APPROVAL KNOPPEN
# ─────────────────────────────────────────────
class StaffApprovalView(discord.ui.View):
    def __init__(self, total, total_views, discord_naam, user_id, betaalmethode, betaalgegevens):
        super().__init__(timeout=None)
        self.total          = total
        self.total_views    = total_views
        self.discord_naam   = discord_naam
        self.user_id        = user_id
        self.betaalmethode  = betaalmethode
        self.betaalgegevens = betaalgegevens

    @discord.ui.button(label="Goedkeuren", style=discord.ButtonStyle.success, custom_id="approve_pay")
    async def approve(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not is_staff(interaction.user):
            await interaction.response.send_message("Alleen staff.", ephemeral=True)
            return

        await interaction.response.defer()
        self.stop()

        # ── Stap 1: Leaderboard bijwerken ────────────────────────────
        add_to_leaderboard(self.discord_naam, self.user_id, self.total_views, self.total)

        # ── Stap 2: DM naar de gebruiker — vraag om betaallink ───────
        try:
            member = (
                interaction.guild.get_member(self.user_id)
                or await interaction.guild.fetch_member(self.user_id)
            )
        except Exception as e:
            log.error(f"Kon member niet ophalen: {e}")
            member = None

        betaallink_ontvangen = False

        if member:
            dm_vraag_embed = discord.Embed(
                title="💰 Je betaalverzoek is goedgekeurd!",
                description=(
                    f"Hey **{self.discord_naam}**! Je clipfarming-verzoek van "
                    f"**EUR {self.total:.2f}** is goedgekeurd. 🎉\n\n"
                    "**Stuur nu je betaallink als reply op dit bericht.**\n"
                    "Stuur een Tikkie-link, PayPal.me-link, of een ander betaalverzoek. "
                    "Wij maken dan het exacte bedrag over.\n\n"
                    "⏳ Je hebt **10 minuten** om te reageren."
                ),
                color=0x00c9a7,
                timestamp=now_utc(),
            )
            dm_vraag_embed.set_footer(text="MONSTERBOT — MonsterTube Clipfarming")

            dm_sent = await send_dm(member, dm_vraag_embed)

            if dm_sent:
                # Wacht op reply van de gebruiker in DMs
                def check(m):
                    return m.author.id == self.user_id and isinstance(m.channel, discord.DMChannel)

                try:
                    reply = await bot.wait_for("message", check=check, timeout=600)  # 10 minuten
                    betaallink = sanitize(reply.content, 300)
                    betaallink_ontvangen = True
                    log.info(f"Betaallink ontvangen van {member}: {betaallink}")
                except asyncio.TimeoutError:
                    timeout_embed = discord.Embed(
                        title="⏰ Tijd verstreken",
                        description=(
                            "Je hebt niet op tijd een betaallink gestuurd.\n"
                            "Neem contact op met staff via een nieuw ticket."
                        ),
                        color=0xff4444,
                    )
                    await send_dm(member, timeout_embed)
                    betaallink = "Geen link ontvangen (timeout)"
                    log.warning(f"Geen betaallink ontvangen van {member} binnen 10 minuten")
            else:
                betaallink = "Kon geen DM sturen — DMs uitgeschakeld"
        else:
            betaallink = "Gebruiker niet gevonden"

        # ── Stap 3: Post naar betaalkanaal ───────────────────────────
        if PAYMENT_CH_ID:
            try:
                pay_ch = interaction.guild.get_channel(PAYMENT_CH_ID)
                if pay_ch:
                    pay_embed = discord.Embed(
                        title="💸 Nieuwe uitbetaling",
                        color=0x00c9a7,
                        timestamp=now_utc(),
                    )
                    pay_embed.add_field(name="Naam",             value=self.discord_naam,         inline=True)
                    pay_embed.add_field(name="Discord",          value=f"<@{self.user_id}>",      inline=True)
                    pay_embed.add_field(name="Bedrag",           value=f"**EUR {self.total:.2f}**", inline=True)
                    pay_embed.add_field(
                        name="Betaallink",
                        value=betaallink if betaallink_ontvangen else f"⚠️ {betaallink}",
                        inline=False,
                    )
                    pay_embed.add_field(name="Goedgekeurd door", value=interaction.user.mention,  inline=False)
                    pay_embed.set_footer(text="Klik op ✅ Betaald als het is overgemaakt.")

                    pay_view = PaymentProcessView(
                        user_id=self.user_id,
                        bedrag=f"EUR {self.total:.2f}",
                        discord_naam=self.discord_naam,
                    )
                    await pay_ch.send(embed=pay_embed, view=pay_view)
            except Exception as e:
                log.error(f"Fout bij sturen naar betaalkanaal: {e}")

        # ── Stap 4: Embed in het ticket — betaling onderweg ──────────
        ticket_embed = discord.Embed(
            title="💸 Betaling is onderweg!",
            description=(
                f"<@{self.user_id}> Je verzoek van **EUR {self.total:.2f}** is goedgekeurd "
                f"en de betaling wordt nu verwerkt.\n\n"
                "**Zodra je het geld hebt ontvangen, klik je op de knop hieronder.**\n"
                "Het ticket sluit daarna automatisch."
            ),
            color=0x00c9a7,
            timestamp=now_utc(),
        )
        ticket_embed.set_footer(text="MONSTERBOT — Klik op de knop zodra je betaling binnen is.")

        try:
            await interaction.channel.send(
                content=f"<@{self.user_id}>",
                embed=ticket_embed,
                view=ConfirmReceiptView(user_id=self.user_id),
            )
        except Exception as e:
            log.error(f"Fout bij sturen ticket-embed na goedkeuring: {e}")

        # ── Stap 5: Log ──────────────────────────────────────────────
        log_embed = discord.Embed(
            title="Betaling goedgekeurd",
            description=(
                f"**Door:** {interaction.user.mention}\n"
                f"**Bedrag:** EUR {self.total:.2f}\n"
                f"**Ontvanger:** <@{self.user_id}>\n"
                f"**Betaallink:** {betaallink}\n"
                f"**Datum:** {now_local().strftime('%d/%m/%Y %H:%M')}"
            ),
            color=0x00c9a7,
            timestamp=now_utc(),
        )
        log_embed.set_footer(text="MONSTERBOT - Ticket Log")
        await send_log(interaction.guild, log_embed)
        log.info(f"Betaling goedgekeurd door {interaction.user} — EUR {self.total:.2f}")

    @discord.ui.button(label="Afwijzen", style=discord.ButtonStyle.danger, custom_id="reject_pay")
    async def reject(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not is_staff(interaction.user):
            await interaction.response.send_message("Alleen staff.", ephemeral=True)
            return
        await interaction.response.send_modal(RejectModal(self.user_id, self.discord_naam))

    @discord.ui.button(label="Nakijken", style=discord.ButtonStyle.secondary, custom_id="review_pay")
    async def review(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not is_staff(interaction.user):
            await interaction.response.send_message("Alleen staff.", ephemeral=True)
            return
        try:
            new_name = f"nakijken-{interaction.channel.name.replace('ticket-', '')}"
            await interaction.channel.edit(name=new_name)
        except Exception as e:
            log.warning(f"Kon kanaal niet hernoemen: {e}")

        embed = discord.Embed(
            title="Moet worden nagekeken",
            description=(
                f"Dit ticket is gemarkeerd als **te nakijken** door {interaction.user.mention}.\n"
                "Staff zal dit zo snel mogelijk controleren."
            ),
            color=0xf0a500,
            timestamp=now_utc(),
        )
        await interaction.response.send_message(embed=embed)

        log_embed = discord.Embed(
            title="Ticket gemarkeerd: Nakijken",
            description=(
                f"**Door:** {interaction.user.mention}\n"
                f"**Kanaal:** {interaction.channel.mention}\n"
                f"**Datum:** {now_local().strftime('%d/%m/%Y %H:%M')}"
            ),
            color=0xf0a500,
            timestamp=now_utc(),
        )
        log_embed.set_footer(text="MONSTERBOT - Ticket Log")
        await send_log(interaction.guild, log_embed)


class RejectModal(discord.ui.Modal, title="Afwijzingsreden"):
    reden = discord.ui.TextInput(
        label="Waarom wordt dit afgewezen?",
        style=discord.TextStyle.paragraph,
        placeholder="Bijv. views niet geverifieerd, verkeerd template...",
        required=True,
        max_length=500,
    )

    def __init__(self, user_id: int, discord_naam: str):
        super().__init__()
        self.user_id      = user_id
        self.discord_naam = discord_naam

    async def on_submit(self, interaction: discord.Interaction):
        reden_clean = sanitize(self.reden.value, 500)

        embed = discord.Embed(
            title="Afgewezen",
            description=f"**Reden:** {reden_clean}\n\nPas je inzending aan en probeer opnieuw.",
            color=0xff4444,
        )
        await interaction.response.send_message(embed=embed)

        # DM naar de gebruiker
        try:
            member = (
                interaction.guild.get_member(self.user_id)
                or await interaction.guild.fetch_member(self.user_id)
            )
            dm_embed = discord.Embed(
                title="❌ Betaalverzoek afgewezen",
                description=(
                    f"Hey **{self.discord_naam}**, helaas is je betaalverzoek afgewezen.\n\n"
                    f"**Reden:** {reden_clean}\n\n"
                    "Pas je inzending aan en maak een nieuw ticket aan."
                ),
                color=0xff4444,
                timestamp=now_utc(),
            )
            dm_embed.set_footer(text="MONSTERBOT — MonsterTube Clipfarming")
            await send_dm(member, dm_embed)
        except Exception as e:
            log.warning(f"Kon member niet ophalen voor DM bij afwijzing: {e}")

        log_embed = discord.Embed(
            title="Betaling afgewezen",
            description=(
                f"**Door:** {interaction.user.mention}\n"
                f"**Reden:** {reden_clean}\n"
                f"**Datum:** {now_local().strftime('%d/%m/%Y %H:%M')}"
            ),
            color=0xff4444,
            timestamp=now_utc(),
        )
        log_embed.set_footer(text="MONSTERBOT - Ticket Log")
        await send_log(interaction.guild, log_embed)
        log.info(f"Betaling afgewezen door {interaction.user}: {reden_clean[:50]}")

# ─────────────────────────────────────────────
#  EXCEL UPLOAD VERWERKEN
# ─────────────────────────────────────────────
@bot.event
async def on_message(message):
    if message.author.bot:
        return

    # DM-berichten worden afgehandeld via wait_for — niet hier verwerken
    if isinstance(message.channel, discord.DMChannel):
        return

    if not message.channel.category or message.channel.category.id != TICKET_CAT_ID:
        await bot.process_commands(message)
        return

    for attachment in message.attachments:
        if not attachment.filename.lower().endswith((".xlsx", ".xls")):
            continue
        try:
            file_bytes = await attachment.read()
        except Exception as e:
            log.error(f"Kon bijlage niet lezen: {e}")
            await message.channel.send("Kon het bestand niet lezen. Probeer opnieuw.")
            continue

        data = parse_submission(file_bytes)
        if not data:
            embed = discord.Embed(
                title="Bestand niet leesbaar",
                description="Gebruik het meegeleverde template en vul de Views-kolom correct in.",
                color=0xff8c00,
            )
            await message.channel.send(embed=embed)
            continue

        topic          = message.channel.topic or ""
        betaalmethode  = "Onbekend"
        betaalgegevens = "Onbekend"
        try:
            parts = topic.split("|")
            if len(parts) >= 2:
                pay_raw = parts[1].strip()
                if ":" in pay_raw:
                    betaalmethode, betaalgegevens = pay_raw.split(":", 1)
                    betaalmethode  = sanitize(betaalmethode, 50)
                    betaalgegevens = sanitize(betaalgegevens, 150)
        except Exception:
            pass

        staff_role = message.guild.get_role(STAFF_ROLE_ID)
        view = StaffApprovalView(
            total=data["total"],
            total_views=data["total_views"],
            discord_naam=data["discord_naam"],
            user_id=message.author.id,
            betaalmethode=betaalmethode,
            betaalgegevens=betaalgegevens,
        )

        try:
            await message.channel.send(
                content=(f"{staff_role.mention} - nieuw verzoek ter beoordeling!" if staff_role else ""),
                embed=build_summary_embed(data, message.author),
                view=view,
            )
        except Exception as e:
            log.error(f"Fout bij sturen summary: {e}")
            continue

        log_embed = discord.Embed(
            title="Excel ingediend",
            description=(
                f"**Gebruiker:** {message.author.mention}\n"
                f"**Naam:** {data['discord_naam']}\n"
                f"**Clips:** {len(data['clips'])}\n"
                f"**Totaal:** EUR {data['total']:.2f}\n"
                f"**Datum:** {now_local().strftime('%d/%m/%Y %H:%M')}"
            ),
            color=0xf0a500,
            timestamp=now_utc(),
        )
        log_embed.set_footer(text="MONSTERBOT - Ticket Log")
        if LOG_CHANNEL_ID:
            try:
                log_ch = message.guild.get_channel(LOG_CHANNEL_ID)
                if log_ch:
                    excel_file = discord.File(
                        io.BytesIO(file_bytes),
                        filename=f"{message.author.name}_{now_local().strftime('%d-%m-%Y_%H-%M')}.xlsx"
                    )
                    await log_ch.send(embed=log_embed, file=excel_file)
            except Exception as e:
                log.error(f"Fout bij log Excel upload: {e}")

        log.info(f"Excel ingediend door {message.author} — EUR {data['total']:.2f}")

    await bot.process_commands(message)

# ─────────────────────────────────────────────
#  LEADERBOARD
# ─────────────────────────────────────────────
async def post_leaderboard(guild, month_key):
    if not LEADERBOARD_CH_ID:
        return
    try:
        channel = guild.get_channel(LEADERBOARD_CH_ID)
        if not channel:
            return
        data = load_data()
        if month_key not in data or not data[month_key]:
            return
        entries     = sorted(data[month_key].values(), key=lambda x: x["views"], reverse=True)[:10]
        year, month = month_key.split("-")
        month_name  = MONTHS_NL[int(month) - 1]
        embed = discord.Embed(
            title=f"Leaderboard — {month_name} {year}",
            color=0xf0a500,
            timestamp=now_utc(),
        )
        medals      = ["1","2","3","4","5","6","7","8","9","10"]
        views_lines = []
        earn_lines  = []
        for i, entry in enumerate(entries):
            m = medals[i] if i < len(medals) else f"{i+1}."
            views_lines.append(f"#{m} **{entry['naam']}** — {entry['views']:,} views")
            earn_lines.append( f"#{m} **{entry['naam']}** — EUR {entry['earnings']:.2f}")
        embed.add_field(name="Views",       value="\n".join(views_lines) or "Geen data", inline=True)
        embed.add_field(name="Verdiensten", value="\n".join(earn_lines)  or "Geen data", inline=True)
        await channel.send(embed=embed)
    except Exception as e:
        log.error(f"post_leaderboard fout: {e}")

@tree.command(name="leaderboard", description="[Staff] Post het leaderboard van de huidige maand.")
async def leaderboard_cmd(interaction: discord.Interaction):
    if not is_staff(interaction.user):
        await interaction.response.send_message("Alleen staff.", ephemeral=True)
        return
    await interaction.response.defer(ephemeral=True)
    await post_leaderboard(interaction.guild, get_month_key())
    await interaction.followup.send("Leaderboard gepost!", ephemeral=True)

@tasks.loop(hours=1)
async def check_monthly_leaderboard():
    try:
        now      = now_local()
        last_day = calendar.monthrange(now.year, now.month)[1]
        if now.day == last_day and now.hour == 22:
            for guild in bot.guilds:
                await post_leaderboard(guild, get_month_key())
    except Exception as e:
        log.error(f"check_monthly_leaderboard fout: {e}")

# ─────────────────────────────────────────────
#  STAFF COMMANDS
# ─────────────────────────────────────────────
@tree.command(name="bot_aan", description="[Staff] Zet de bot aan voor alle gebruikers.")
async def bot_aan(interaction: discord.Interaction):
    if not is_staff(interaction.user):
        await interaction.response.send_message("Alleen staff.", ephemeral=True)
        return
    state = load_state(); state["bot_enabled"] = True; save_state(state)
    await interaction.response.send_message("Bot ingeschakeld.")

@tree.command(name="bot_uit", description="[Staff] Zet de bot uit voor alle gebruikers.")
async def bot_uit(interaction: discord.Interaction):
    if not is_staff(interaction.user):
        await interaction.response.send_message("Alleen staff.", ephemeral=True)
        return
    state = load_state(); state["bot_enabled"] = False; save_state(state)
    await interaction.response.send_message("Bot uitgeschakeld.")

@tree.command(name="bot_status", description="[Staff] Bekijk de huidige status van de bot.")
async def bot_status(interaction: discord.Interaction):
    if not is_staff(interaction.user):
        await interaction.response.send_message("Alleen staff.", ephemeral=True)
        return
    state = load_state(); enabled = state.get("bot_enabled", True)
    await interaction.response.send_message(
        f"Status: {'OPEN' if is_bot_open() else 'GESLOTEN'} (Handmatig: {'AAN' if enabled else 'UIT'})",
        ephemeral=True,
    )

@tree.command(name="betaald", description="[Staff] Markeer ticket als betaald en sluit het.")
@app_commands.describe(bedrag="Uitbetaald bedrag in euros")
async def betaald(interaction: discord.Interaction, bedrag: str = ""):
    if not is_staff(interaction.user):
        await interaction.response.send_message("Alleen staff.", ephemeral=True)
        return
    if not interaction.channel.category or interaction.channel.category.id != TICKET_CAT_ID:
        await interaction.response.send_message("Geen ticket-kanaal.", ephemeral=True)
        return
    await interaction.response.send_message(f"Betaald: EUR {bedrag}. Sluiten in 10s...")
    await asyncio.sleep(10)
    await interaction.channel.delete()

@tree.command(name="afwijzen", description="[Staff] Wijs een betalingsverzoek af met reden.")
@app_commands.describe(reden="Reden voor afwijzing")
async def afwijzen(interaction: discord.Interaction, reden: str = "Geen reden"):
    if not is_staff(interaction.user):
        await interaction.response.send_message("Alleen staff.", ephemeral=True)
        return
    if not interaction.channel.category or interaction.channel.category.id != TICKET_CAT_ID:
        await interaction.response.send_message("Geen ticket-kanaal.", ephemeral=True)
        return
    await interaction.response.send_message(f"Afgewezen: {reden}")

# ─────────────────────────────────────────────
#  READY
# ─────────────────────────────────────────────
@bot.event
async def on_ready():
    log.info(f"Ingelogd als {bot.user}")
    # Registreer persistente views zodat knoppen werken na herstart
    bot.add_view(TicketButtonView())
    bot.add_view(TicketCloseView(owner_id=0))
    bot.add_view(ConfirmReceiptView(user_id=0))
    try:
        synced = await tree.sync()
        log.info(f"{len(synced)} slash commands globaal gesynchroniseerd")
    except Exception as e:
        log.error(f"Sync mislukt: {e}")
    check_monthly_leaderboard.start()
    # Post ticket-knop in het juiste kanaal
    for guild in bot.guilds:
        await post_ticket_button(guild)

bot.run(TOKEN, log_handler=None)
